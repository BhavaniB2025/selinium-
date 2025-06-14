import openpyxl

filepath = "C:\\Users\\KATWA\\Desktop\\orangehrm_ddt.xlsx"
my_workbook = openpyxl.load_workbook(filepath)
my_active_sheet = my_workbook["Sheet1"]

my_active_sheet.cell(3, 3).value="pass"
my_active_sheet.cell(2, 3).value="fail"
my_active_sheet.cell(4, 3).value="pass"
my_workbook.save(filepath)
