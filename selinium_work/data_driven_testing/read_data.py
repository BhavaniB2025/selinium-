'''
Created on 11 Jun 2025

@author: KATWA
'''
import openpyxl
filepath = "C:\\Users\\KATWA\\Desktop\\orangehrm_ddt.xlsx"
my_workbook = openpyxl.load_workbook(filepath)
my_active_sheet = my_workbook["Sheet1"]   #get the sheet by name
#my_active_sheet = my_workbook.active  #get the currently active sheet
total_rows = my_active_sheet.max_row
print(total_rows)

total_columns = my_active_sheet.max_column
print(total_columns)

print(my_active_sheet.cell(1, 1).value)
print(my_active_sheet.cell(3, 1).value)
print(my_active_sheet.cell(3, 1).value)
print(my_active_sheet.cell(3, 2).value)

for i in range(2, total_rows+1):
    username = my_active_sheet.cell(1, 1).value  #i in place 1
    password = my_active_sheet.cell(1, 2).value  #i in place 1
    print(username, password)

username = my_active_sheet.cell(3, 1).value
password = my_active_sheet.cell(3, 2).value
print(username, password)

username = my_active_sheet.cell(4, 1).value
password = my_active_sheet.cell(4, 2).value
print(username, password)

username = my_active_sheet.cell(2, 2).value
password = my_active_sheet.cell(3, 2).value
print(username, password)